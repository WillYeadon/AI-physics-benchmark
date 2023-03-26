import openai
import openpyxl
from openpyxl import load_workbook

def get_answer_old(prompt, engine_name):
    response = openai.Completion.create(
        engine=engine_name,
        prompt=prompt,
        max_tokens=50,
        n=1,
        stop=None,
        temperature=0.5,
    )
    return response.choices[0].text.strip()

def get_answer(prompt, engine_name):
    response = openai.ChatCompletion.create(
        model=engine_name,
        messages=[
          {"role": "system", "content": "You are a helpful assistant."},
          {"role": "user", "content": prompt},
      ]
    )
    return response['choices'][0]['message']['content']


# Set up the OpenAI API
openai.api_key = ""

# Load the Excel workbook
workbook = load_workbook("Question-Answer-Bank.xlsx")
worksheet = workbook.active

# Define the prompt prefix and suffix
prompt_prefix = "Answer the following question as correctly and concisely as possible ideally using numbers alone including units where appropriate: '"
prompt_suffix = "'? Please answer below."

# Iterate through the rows, starting from row 2 (skipping the header)
for row in range(2, worksheet.max_row + 1):
    question = worksheet.cell(row=row, column=6).value  # Column 6 is the 'Question' column

    if question is not None:
        prompt = prompt_prefix + question + prompt_suffix

        # Get answers from the engines
        answer_davinci_002 = get_answer_old(prompt, "text-davinci-002")
        answer_gpt_35_turbo = get_answer_old(prompt, "text-davinci-002")
#        answer_gpt_4 = get_answer(prompt, "text-davinci-002") 

        ###############
        # CHARGES!
        ###############
        #answer_gpt_35_turbo = get_answer(prompt, "gpt-3.5-turbo") 
        #answer_gpt_4 = get_answer(prompt, "gpt-3.5-turbo") 

        # Store the answers in the corresponding columns
        worksheet.cell(row=row, column=8).value = answer_davinci_002
        worksheet.cell(row=row, column=10).value = answer_gpt_35_turbo
#        worksheet.cell(row=row, column=12).value = answer_gpt_4

# Save the updated workbook
workbook.save("Question-Answer-Bank-Updated.xlsx")
